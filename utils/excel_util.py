import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from flask import send_file, Response


def sheet_Tupla(path):
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active
    result_list = []
    for value in sheet.iter_rows(values_only=True):
        result_list.append(value)
    result_list.pop(0)
    return result_list

def create_workbook(data_list, headers, filename):
    #print (data_list)
    #keys = validate_dict.keys()
    #print (keys)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for index, item in enumerate(data_list):
        #print(index, item)
        ws1 = wb.create_sheet("Data"+str(index+1))
        for row in item:
            ws1.append(row)
    #for element in keys:
    #    print (validate_dict[element])
    #    data_val = DataValidation(type="list",formula1=validate_dict[element]) #You can change =$A:$A with a smaller range like =A1:A9
    #    ws.add_data_validation(data_val)
    #    for num in range(2, max_rows + 2):
    #        data_val.add(ws[element + str(num)]) #If you go to the cell B1 you will find a drop down list with all the values from the column A
    return Response(
        save_virtual_workbook(wb),
        headers={
            'Content-Disposition': 'attachment; filename='+filename+'.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )
def map_to_option(list_):
    string_list = ""
    for element in list_:
        string_list = string_list + str(element) + ","
    return string_list[:-1]


#print (create_workbook({'A': '"Dog,Cat,Bat"', 'C': '"wea1, wea2, wea3"'}))


